from rest_framework.mixins import ListModelMixin, CreateModelMixin, RetrieveModelMixin, UpdateModelMixin, DestroyModelMixin
from extensions.paginations import LimitedPagination, InfinitePagination
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models.deletion import ProtectedError
from rest_framework.viewsets import GenericViewSet
from extensions.exceptions import ValidationError
from django.http.response import HttpResponse
from openpyxl import Workbook, load_workbook
from rest_framework.viewsets import ViewSet


class FunctionViewSet(ViewSet):
    """功能视图"""

    @property
    def team(self):
        return self.request.user.team

    @property
    def user(self):
        return self.request.user

    @property
    def context(self):
        return {'request': self.request, 'format': self.format_kwarg, 'view': self}


class BaseViewSet(GenericViewSet):
    pagination_class = LimitedPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    ordering_fields = ['id']
    ordering = ['-id']
    select_related_fields = []
    prefetch_related_fields = []

    @property
    def team(self):
        return self.request.user.team

    @property
    def user(self):
        return self.request.user

    @property
    def context(self):
        return self.get_serializer_context()

    def get_queryset(self):
        queryset = super().get_queryset().filter(team=self.team)
        queryset = queryset.select_related(*self.select_related_fields)
        queryset = queryset.prefetch_related(*self.prefetch_related_fields)
        return queryset


class ModelViewSet(BaseViewSet, ListModelMixin, CreateModelMixin,
                   RetrieveModelMixin, UpdateModelMixin, DestroyModelMixin):
    """模型视图"""


class PersonalViewSet(BaseViewSet):
    """个人试图"""

    def get_queryset(self):
        return super().get_queryset().filter(creator=self.user)


class LimitedOptionViewSet(BaseViewSet, ListModelMixin):
    """有限选项视图"""


class InfiniteOptionViewSet(BaseViewSet, ListModelMixin):
    """无限选项视图"""

    pagination_class = InfinitePagination


class ReadOnlyMixin(RetrieveModelMixin, ListModelMixin):
    """只读"""


class DataProtectMixin:
    """数据保护"""

    def perform_destroy(self, instance):
        try:
            instance.delete()
        except ProtectedError:
            raise ValidationError('已被数据引用, 无法删除')


class ExportMixin:
    """导出"""

    def get_export_response(self, serializer_class, data=None):
        """获取导出Excel文件响应

        Args:
            serializer_class (BaseSerializer): 序列化类
            data (list): 数据

            serializer (BaseSerializer): 序列化器
            field_items (List): 字段属性

        Returns:
            HttpResponse: 文件响应
        """

        workbook = Workbook()
        work_sheet = workbook.active

        if data:
            results = data
        else:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = serializer_class(instance=queryset, many=True)
            results = serializer.data
        field_items = serializer_class().get_fields().items()

        # 创建表头
        work_sheet.cell(row=1, column=1, value='序号')
        for index, (field_name, field_class) in enumerate(field_items, start=2):
            work_sheet.cell(row=1, column=index, value=field_class.label)

        # 填充数据
        for row, item in enumerate(results, start=2):
            work_sheet.cell(row=row, column=1, value=row - 1)

            for column, (field_name, field_class) in enumerate(field_items, start=2):
                work_sheet.cell(row=row, column=column, value=item.get(field_name, ''))

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response


class ImportMixin:
    """导入"""

    def get_template_response(self, serializer_class):
        """获取Excel模板响应

        Args:
            serializer_class (BaseSerializer): 序列化类

        Returns:
            HttpResponse: 文件响应
        """

        workbook = Workbook()
        work_sheet = workbook.active

        # 创建表头
        field_items = serializer_class().get_fields().items()
        for index, (_, field_class) in enumerate(field_items, start=1):
            work_sheet.cell(row=1, column=index, value=field_class.label)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response

    def load_data(self, file, serializer_class):
        workbook = load_workbook(file)
        work_sheet = workbook.active

        field_items = serializer_class().get_fields().items()
        for row in range(2, work_sheet.max_row + 1):
            data = {}

            for column, (field_name, field_class) in enumerate(field_items):
                if work_sheet[1][column].value != field_class.label:
                    raise ValidationError('格式错误')

                if work_sheet[row][column].value is not None:
                    data[field_name] = work_sheet[row][column].value

            else:
                serializer = serializer_class(data=data, context=self.context)
                serializer.is_valid(raise_exception=True)
                yield serializer


__all__ = [
    'FunctionViewSet', 'BaseViewSet', 'ModelViewSet', 'PersonalViewSet',
    'LimitedOptionViewSet', 'InfiniteOptionViewSet',
    'ReadOnlyMixin', 'DataProtectMixin', 'ExportMixin', 'ImportMixin',
    'ListModelMixin', 'CreateModelMixin', 'RetrieveModelMixin', 'UpdateModelMixin', 'DestroyModelMixin',
    'DjangoFilterBackend', 'SearchFilter', 'OrderingFilter',
]
